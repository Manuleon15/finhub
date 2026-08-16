"""Tests del importador de Excel del Portfolio Tracker."""

import io

import openpyxl

from app.modules.portfolio.importer import parse_portfolio_excel


def _build_test_workbook() -> bytes:
    """Construye un .xlsx mínimo en memoria con el mismo formato que el
    Excel real (hoja DATOS INVERSIONES + hoja MOVIMIENTOS 2025)."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "DATOS INVERSIONES"
    ws.append(
        [
            "NOMBRE DEL ACTIVO",
            "TICKER",
            "CANTIDAD",
            "PRECIO DE COMPRA",
            "PRECIO ACTUAL",
            "BETA",
            "PRECIO OBJETIVO",
        ]
    )
    ws.append(["#VALUE!", "MSFT", 10.0, 400.0, 500.0, 1.1, 450.0])
    ws.append(["#VALUE!", "AAPL", 5.0, 150.0, 180.0, 1.2, 200.0])
    ws.append([None, None, None, None, None, None, None])  # fila vacía, debe ignorarse

    ws2 = wb.create_sheet("MOVIMIENTOS 2025")
    ws2.append(["Ventas 2025", None])
    ws2.append([-16.0, "SL"])
    ws2.append([77.0, "FTNT"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_positions():
    data = _build_test_workbook()
    result = parse_portfolio_excel(data)

    assert result.get("error") is None
    assert len(result["positions"]) == 2

    msft = next(p for p in result["positions"] if p["ticker"] == "MSFT")
    assert msft["quantity"] == 10.0
    assert msft["avg_price"] == 400.0
    assert msft["current_price"] == 500.0
    assert msft["target_price"] == 450.0
    # La columna NOMBRE DEL ACTIVO viene rota (#VALUE!) -> se descarta, no se importa basura
    assert msft["name"] is None


def test_parse_transactions():
    data = _build_test_workbook()
    result = parse_portfolio_excel(data)

    assert len(result["transactions"]) == 2
    sl = next(t for t in result["transactions"] if t["ticker"] == "SL")
    assert sl["realized_pl"] == -16.0
    assert sl["tx_type"] == "sell"


def test_missing_ticker_column_returns_warning():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATOS INVERSIONES"
    ws.append(["ALGO", "OTRA COSA"])
    ws.append([1, 2])

    buf = io.BytesIO()
    wb.save(buf)

    result = parse_portfolio_excel(buf.getvalue())
    assert result["positions"] == []
    assert len(result["warnings"]) == 1
