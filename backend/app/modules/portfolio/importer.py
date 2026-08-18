"""Importador completo del Excel de portfolio (5 hojas).

Hojas que lee:
- DATOS INVERSIONES: posiciones (19 filas)
- DIVERSIFICACIÓN: sectorización por ticker
- RESUMEN RETORNOS MENSUALES: retornos mensuales vs SP500 (histórico)
- DASHBOARD: KPIs agregados (opcional, para referencia)
- MOVIMIENTOS <año>: transacciones de venta

Todos los campos se leen por nombre de columna, no por posición.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Dict, List, Optional

import openpyxl

logger = logging.getLogger("finhub.portfolio.importer")

POSITIONS_SHEET = ["DATOS INVERSIONES", "POSICIONES", "PORTFOLIO"]
DIVERSIFICATION_SHEET = ["DIVERSIFICACIÓN", "DIVERSIFICACION", "SECTORES"]
MONTHLY_SHEET = ["RESUMEN RETORNOS MENSUALES", "RETORNOS", "HISTORICO"]
DASHBOARD_SHEET = ["DASHBOARD"]
TRANSACTIONS_PREFIX = "MOVIMIENTOS"

# Mapeo Excel -> modelo Position
POSITION_COLUMN_MAP = {
    "TICKER": "ticker",
    "NOMBRE DEL ACTIVO": "name",
    "CANTIDAD": "quantity",
    "PRECIO DE COMPRA": "avg_price",
    "PRECIO ACTUAL": "current_price",
    "PRECIO OBJETIVO": "target_price",
    "BETA": "beta",
    "DIVIDENDO X ACCION": "dividend_per_share",
    "P/G REALIZADAS": "realized_pl",
}

# Meses: Año '25, etc.
MONTH_PATTERN = re.compile(r"^(Ene|Feb|Mar|Abr|May|Jun|Jul|Ago|Sep|Oct|Nov|Dic)\s*'(\d{2})$")
MONTH_ORDER = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
               "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    for a, b in {"Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ñ": "N"}.items():
        s = s.replace(a, b)
    return s


def _find_sheet(wb: openpyxl.Workbook, candidates: List[str]) -> Optional[str]:
    norm_names = {_norm(n): n for n in wb.sheetnames}
    for c in candidates:
        if _norm(c) in norm_names:
            return norm_names[_norm(c)]
    return None


def _num(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, str):
            v = v.replace("%", "").replace(",", ".").strip()
            if not v:
                return None
        f = float(v)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def parse_positions_sheet(wb: openpyxl.Workbook, warnings: List[str]) -> List[Dict[str, Any]]:
    sheet = _find_sheet(wb, POSITIONS_SHEET)
    if not sheet:
        warnings.append(f"No encontré la hoja de posiciones (probé: {POSITIONS_SHEET})")
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = rows[0]
    idx: Dict[str, int] = {}
    for i, name in enumerate(header):
        norm = _norm(name)
        for excel_col, field in POSITION_COLUMN_MAP.items():
            if _norm(excel_col) == norm:
                idx[field] = i
    if "ticker" not in idx:
        warnings.append(f"La hoja '{sheet}' no tiene columna TICKER")
        return []

    out = []
    for row in rows[1:]:
        ticker = row[idx["ticker"]] if idx["ticker"] < len(row) else None
        if not ticker or not isinstance(ticker, str) or ticker.startswith("#"):
            continue
        pos: Dict[str, Any] = {"ticker": ticker.strip().upper()}
        for field, col in idx.items():
            if field == "ticker" or col >= len(row):
                continue
            v = row[col]
            if field == "name":
                pos[field] = v if isinstance(v, str) and not v.startswith("#") else None
            else:
                pos[field] = _num(v)
        if pos.get("quantity") is None:
            warnings.append(f"{pos['ticker']}: sin CANTIDAD, se pone a 0.")
            pos["quantity"] = 0.0
        out.append(pos)
    return out


def parse_diversification_sheet(wb: openpyxl.Workbook) -> Dict[str, str]:
    """Lee DIVERSIFICACIÓN → {ticker: sector}"""
    sheet = _find_sheet(wb, DIVERSIFICATION_SHEET)
    if not sheet:
        return {}
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 3:
            continue
        ticker = row[1]  # Columna B
        sector = row[3]  # Columna D (INDUSTRIA ESPAÑOL)
        if ticker and isinstance(ticker, str) and sector and isinstance(sector, str):
            out[ticker.strip().upper()] = sector.strip()
    return out


def parse_monthly_sheet(wb: openpyxl.Workbook, warnings: List[str]) -> List[Dict[str, Any]]:
    """Lee RESUMEN RETORNOS MENSUALES.

    Estructura: filas son conceptos (retorno portfolio, retorno sp500, valor),
    columnas son meses (Año '25, Ene '26...). Detecta los dos años y genera snapshots.
    """
    sheet = _find_sheet(wb, MONTHLY_SHEET)
    if not sheet:
        warnings.append("No encontré la hoja de retornos mensuales")
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = rows[0]
    month_cols: List[tuple[int, int, int]] = []  # (col_idx, year, month)

    for col_idx, name in enumerate(header):
        if not isinstance(name, str):
            continue
        m = MONTH_PATTERN.match(name.strip())
        if m:
            mon_name, yr = m.group(1).lower(), m.group(2)
            if mon_name in MONTH_ORDER:
                month_cols.append((col_idx, 2000 + int(yr), MONTH_ORDER[mon_name]))

    if not month_cols:
        warnings.append("La hoja de retornos no tiene meses detectables (formato 'Ene 25').")
        return []

    # Encontrar filas clave
    portfolio_value = None
    portfolio_twr = None
    sp500_twr = None

    for row in rows:
        if len(row) < 2:
            continue
        label = _norm(row[0])
        if "VALOR PORTFOLIO" in label:
            portfolio_value = row
        elif "RETORNO MENSUAL PORTFOLIO" in label or "RETORNO MENSUAL CARTERA" in label:
            portfolio_twr = row
        elif "RETORNO MENSUAL SP500" in label or "RETORNO MENSUAL S&P500" in label:
            sp500_twr = row

    snapshots = []
    for col_idx, year, month in month_cols:
        if portfolio_value is None or col_idx >= len(portfolio_value):
            continue
        pv = _num(portfolio_value[col_idx]) if portfolio_value else None
        port = _num(portfolio_twr[col_idx]) if portfolio_twr else None
        sp = _num(sp500_twr[col_idx]) if sp500_twr else None

        snapshots.append({
            "year": year,
            "month": month,
            "portfolio_value": pv,
            "portfolio_twr_month": port or 0.0,
            "sp500_twr_month": sp or 0.0,
        })

    return snapshots


def parse_transactions(wb: openpyxl.Workbook, warnings: List[str]) -> List[Dict[str, Any]]:
    out = []
    for name in wb.sheetnames:
        if not _norm(name).startswith(TRANSACTIONS_PREFIX):
            continue
        ws = wb[name]
        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if len(row) < 2:
                continue
            pl, ticker = row[0], row[1]
            if not isinstance(ticker, str) or not ticker.strip():
                continue
            if not _is_num(pl):
                continue
            out.append({
                "ticker": ticker.strip().upper(),
                "tx_type": "sell",
                "realized_pl": float(pl),
                "quantity": None,
                "price": None,
                "date": None,
                "notes": f"De hoja '{name}'",
            })
    return out


def _is_num(v: Any) -> bool:
    return isinstance(v, (int, float))


def parse_portfolio_excel(file_bytes: bytes) -> Dict[str, Any]:
    warnings: List[str] = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"error": f"No se pudo leer el Excel: {e}", "positions": [], "transactions": [], "warnings": []}

    positions = parse_positions_sheet(wb, warnings)
    sectors = parse_diversification_sheet(wb)
    snapshots = parse_monthly_sheet(wb, warnings)
    transactions = parse_transactions(wb, warnings)

    # Asignar sector a cada posición
    for pos in positions:
        pos["sector"] = sectors.get(pos["ticker"])

    logger.info(f"Excel: {len(positions)} pos, {len(transactions)} tx, {len(snapshots)} snapshots, {len(sectors)} sectores, {len(warnings)} avisos")

    return {
        "positions": positions,
        "transactions": transactions,
        "monthly_snapshots": snapshots,
        "warnings": warnings,
    }
